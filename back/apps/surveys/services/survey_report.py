from django.db.models import Q
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from apps.docs.utils.excel_utils import set_font_bold, set_cell_border
from apps.edu.consts.student_group.type import EDU, INFO
from apps.edu.services.student_group import student_group_service
from apps.surveys.selectors.student_answer import student_answer_orm
from apps.surveys.services.survey import survey_service


class SurveyReportService:
    """
    Класс для генерации отчета по результатам опроса
    """

    # Последняя занятая строка в листе
    last_row = 0

    def __init__(self, report_parameters: dict):
        """
        Инициализация класса - фиксаци ID опроса и параметров отчета
        :param report_parameters: параметры отчета
        """
        self.survey = survey_service.get_survey('object_id', report_parameters.get('survey_id'))
        self.report_type = report_parameters.get('type')
        self.service_type = report_parameters.get('service_type')
        self.group = student_group_service.get_student_group('object_id', report_parameters.get('group_id'))
        self.start_period = report_parameters.get('start_period')
        self.end_period = report_parameters.get('end_period')

    def _fill_main_info(self, wb: Workbook) -> Worksheet:
        """
        Заполнение основной информации отчета
        :param wb: Рабочая книга Excel
        :return:
        """
        worksheet = wb.active
        worksheet.title = 'Результаты опроса'
        cell = worksheet.cell(row=1, column=1)
        set_font_bold(cell)
        set_cell_border(cell)
        cell.value = 'Опрос: '
        cell = worksheet.cell(row=1, column=2)
        set_cell_border(cell)
        cell.value = self.survey.description[:100]
        return worksheet

    def _fill_detail_info(self, ws: Worksheet):
        """
        Заполнение информации по детализации отчета
        :param ws: Лист книги Excel
        :return:
        """
        cell = ws.cell(row=2, column=1)
        set_font_bold(cell)
        set_cell_border(cell)
        cell.value = 'Детализация: '
        cell = ws.cell(row=2, column=2)
        set_cell_border(cell)
        if self.report_type == 'group':
            cell.value = 'По группе'
            cell = ws.cell(row=3, column=1)
            set_font_bold(cell)
            set_cell_border(cell)
            cell.value = 'Код группы:'
            cell = ws.cell(row=3, column=2)
            set_cell_border(cell)
            cell.value = self.group.code
        else:
            if self.report_type == 'service_type':
                cell.value = 'По типу услуг'
                cell = ws.cell(row=3, column=1)
                set_font_bold(cell)
                set_cell_border(cell)
                cell.value = 'Тип услуг:'
                cell = ws.cell(row=3, column=2)
                set_cell_border(cell)
                cell.value = 'Образовательные услуги' if self.service_type == 'edu' else \
                    'Информационно-консультационные услуги'
            else:
                cell.value = 'Все данные'
            cell = ws.cell(row=4, column=1)
            set_font_bold(cell)
            set_cell_border(cell)
            cell.value = 'Период данных: '
            cell = ws.cell(row=4, column=2)
            set_cell_border(cell)
            cell.value = f'{self.start_period} - {self.end_period}'

    def _fill_answers_table(self, ws: Worksheet) -> dict:
        """
        Заполнение таблицы с ответами
        :param ws: Лист книги Excel
        :return: словарь со статистикой
        """
        answers_records = student_answer_orm.get_filter_records(filter_by=dict(survey_id=self.survey.object_id))
        if self.report_type == 'group':
            answers_records = answers_records.filter(group_code=self.group.code)
        else:
            if self.report_type == 'service_type':
                if self.service_type == 'edu':
                    answers_records = answers_records.filter(group_type=EDU)
                else:
                    answers_records = answers_records.filter(group_type=INFO)
                answers_records = answers_records.filter(
                    Q(date_create__gte=self.start_period) & Q(date_create__lte=self.end_period)
                )
        questions = [rec.get('question') for rec in answers_records.values('question').distinct()]
        statistics = {}
        for column_index, question in enumerate(questions, start=1):
            statistics[question] = {}
            cell = ws.cell(row=5, column=column_index)
            set_font_bold(cell)
            set_cell_border(cell)
            cell.value = question
            question_answers = answers_records.filter(question=question).order_by('-object_id')
            for row_index, q_answer in enumerate(question_answers, start=1):
                cell = ws.cell(row=5 + row_index, column=column_index)
                set_cell_border(cell)
                cell.value = q_answer.answer
                # Количество таких же ответов на вопрос
                answer_count = question_answers.filter(answer=q_answer.answer).count()
                # Запись процентного соотношения в статистику
                statistics[question][q_answer.answer] = answer_count / question_answers.count()
                self.last_row = 5 + row_index
        return statistics

    @staticmethod
    def _draw_statistic_chart(ws: Worksheet, question: str, answer_dict: dict, row_index: int):
        """
        Отрисовка гистограммы по статистике ответов на вопрос
        :param ws: Лист книги Excel
        :param question: формулировка вопроса
        :param answer_dict: словарь с ответами и процентом
        :param row_index: индекс строки для отрисовки гистограммы
        :return:
        """
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = 'Статистика ответов на вопрос "' + question + '"'
        chart.y_axis.number_format = '0%'
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1
        percent = Reference(ws, min_col=2, min_row=row_index + 1, max_row=row_index + len(answer_dict) + 1)
        answers = Reference(ws, min_col=1, min_row=row_index + 1, max_row=row_index + len(answer_dict) + 1)
        chart.add_data(percent, titles_from_data=True)
        chart.set_categories(answers)
        chart.width = 20
        chart.height = 7
        ws.add_chart(chart, f"D{row_index}")

    def _fill_statistics(self, ws: Worksheet, statistics: dict):
        """
        Заполнение таблиц со статистикой по вопросам
        :param ws: Лист книги Excel
        :param statistics: словарь со статистикой ответов на вопросы
        :return:
        """
        # Пропускаем строчку и начинаем записывать статистику в следующую
        self.last_row += 2
        for question, answers in statistics.items():
            row_index = self.last_row
            ws.row_dimensions[row_index].height = 30
            ws.merge_cells(f'{get_column_letter(1)}{row_index}:{get_column_letter(2)}{row_index}')
            cell = ws.cell(row=row_index, column=1)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            set_cell_border(cell)
            cell.font = Font(italic=True)
            cell.value = 'Статистика ответов на вопрос "' + question + '"'
            cell = ws.cell(row=row_index + 1, column=1)
            set_cell_border(cell)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.font = Font(italic=True)
            cell.value = "Ответ"
            cell = ws.cell(row=row_index + 1, column=2)
            set_cell_border(cell)
            cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            cell.font = Font(italic=True)
            cell.value = "Процент"
            for answer_index, (answer, percent) in enumerate(answers.items(), start=1):
                cell = ws.cell(row=row_index + answer_index, column=1)
                set_cell_border(cell)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
                cell.value = answer
                cell = ws.cell(row=row_index + answer_index, column=2)
                set_cell_border(cell)
                cell.number_format = '0%'
                cell.value = percent if percent != 1.0 else 1
            self._draw_statistic_chart(ws, question, answers, self.last_row)
            self.last_row += len(answers) + 2

    def _create_workbook(self) -> Workbook:
        """
        Генерация книги Excel
        :return:
        """
        workbook = Workbook()
        ws = self._fill_main_info(workbook)
        self._fill_detail_info(ws)
        stats = self._fill_answers_table(ws)
        self._fill_statistics(ws, stats)
        for i in range(1, ws.max_column + 1):
            letter = get_column_letter(i)
            ws.column_dimensions[letter].width = 35
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        return workbook

    def generate_report(self) -> Workbook:
        """Генерация файла отчета"""
        return self._create_workbook()
