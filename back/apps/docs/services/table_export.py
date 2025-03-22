from django.db.models import QuerySet, ManyToOneRel, ManyToManyRel, OneToOneRel, CharField, TextField, IntegerField, \
    PositiveIntegerField, EmailField, DateField, DateTimeField, ForeignKey, BooleanField
from django.db.models.fields.related import ManyToManyField
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from apps.commons.models import BaseTable
from apps.docs.utils.excel_utils import set_cell_value, set_model_fields_name_to_sheet, set_cells_width


class TableExport:
    """
    Класс для выгрузки данных из таблиц в Excel
    """

    # Список полей, исключенные из фиксации в Excel файле
    _exclude_fields = [
        'object_id',
        'django_user',
        'date_create',
        'old_id',
        'kug_edit',
        'survey_show'
    ]

    def __init__(self, model: BaseTable, queryset: QuerySet):
        """
        Инициализация класса - установка модели (таблицы БД) и queryset с данными выгрузки
        :param model: Модель БД
        :param queryset: Данные
        """
        self._records = None
        self._fields = []
        self.model = model
        self.queryset = queryset

    def _set_fields(self):
        """
        Установка столбцов (поля таблицы БД)
        :return:
        """
        for field in self.model._meta.get_fields():
            if field.name not in self._exclude_fields\
                    and not isinstance(field, (ManyToOneRel, ManyToManyRel, OneToOneRel)):
                self._fields.append(field)

    def _create_records_generator(self):
        """
        Создание генератора записей
        :return:
        """
        self._records = (rec for rec in self.queryset)

    def _generate_wb(self) -> Workbook:
        """
        Генерация workbook
        :return:
        """
        wb = Workbook()
        ws = wb.create_sheet(self.model._meta.verbose_name_plural, -1)
        set_model_fields_name_to_sheet(ws, self._fields)
        self._add_content_to_sheet(sheet=ws)
        set_cells_width(ws, 50)
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        return wb

    def _add_content_to_sheet(self, sheet):
        """
        Добавление данных на лист книги
        :param sheet: лист книги Excel
        :return:
        """
        for row, record in enumerate(self._records, start=2):
            for col, field in enumerate(self._fields, start=1):
                cell = sheet.cell(row=row, column=col)
                set_cell_value(
                    cell,
                    self._get_value_by_field(record, field),
                    True,
                    False,
                    True
                )

    @staticmethod
    def _get_value_by_field(record, field):
        """
        Получение значения поля записи
        :param record: Запись с данными
        :param field: Поле таблицы в БД
        :return: значение
        """
        try:
            value = getattr(record, field.name)
        except AttributeError:
            return '#ERROR#'
        if value is None:
            return '-'
        if isinstance(field, (CharField, TextField, IntegerField, PositiveIntegerField, EmailField)):
            if isinstance(field, CharField):
                if field.choices:
                    try:
                        return dict(field.choices)[value]
                    except KeyError:
                        pass
            return value
        if isinstance(field, BooleanField):
            if field.name == 'sex':
                return 'М' if value else 'Ж'
            return '+' if value else '-'
        if isinstance(field, DateField):
            return value.strftime('%d.%m.%Y')
        if isinstance(field, DateTimeField):
            return value.strftime('%d.%m.%Y %H:%M')
        if isinstance(field, ForeignKey):
            return str(value)
        if isinstance(field, ManyToManyField):
            objects = ''
            for obj in value.all():
                objects += f'{str(obj)}; '
            return objects[:-2]
        return f'Неизвестный тип - {repr(field)}'

    def get_excel(self) -> HttpResponse:
        """
        Получение excel файла с данными
        :return:
        """
        self._set_fields()
        self._create_records_generator()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{self.model._meta.verbose_name_plural}.xlsx"'
        wb = self._generate_wb()
        wb.save(response)

        return response
