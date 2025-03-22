from openpyxl.styles import Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def set_font_bold(cell):
    """Установка в ячейке жирный шрифт текста"""
    cell.font = Font(bold=True)


def set_cell_alignment_center(cell):
    """Установить выравнивание текста в ячейке по центру по вертикали и горизонтали"""
    cell.alignment = Alignment(
        wrap_text=True,
        horizontal='center',
        vertical='center'
    )


def set_cell_border(cell):
    """Отрисовка границ ячейки"""
    thin = Side(border_style="thin", color="000000")
    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def set_cell_value(cell, value, align_center: bool = False, font_bold: bool = False, border: bool = False):
    """
    Установка значения в ячейки, выравнивания, установка жирного шрифта, отрисовка границы ячейки
    :param cell: ячейка листа
    :param value: значение
    :param align_center: выравнивание по центру
    :param font_bold: жирный шрифт
    :param border: отрисовка границы
    :return:
    """
    cell.value = value
    if align_center:
        set_cell_alignment_center(cell)
    if font_bold:
        set_font_bold(cell)
    if border:
        set_cell_border(cell)


def set_cells_width(ws: Worksheet, width: int):
    """
    Установка ширины ячеек на всем листе книги Excel
    :param ws:
    :param width:
    :return:
    """
    for index, col in enumerate(ws, start=1):
        ws.column_dimensions[get_column_letter(index)].width = 50


def set_model_fields_name_to_sheet(sheet, fields: list):
    """
    Добавление наименований столбцов полей модели БД в первую строку листа
    :param sheet: лист книги Excel
    :param fields: Список полей модели БД
    :return:
    """
    for index, field in enumerate(fields, start=1):
        cell = sheet.cell(row=1, column=index)
        try:
            value = field.verbose_name
        except AttributeError:
            value = field.name
        set_cell_value(cell, value, True, True, True)


