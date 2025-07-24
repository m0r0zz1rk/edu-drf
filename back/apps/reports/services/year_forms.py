import datetime
from typing import Union

from django.db.models import QuerySet
from openpyxl import Workbook

from apps.applications.consts.application_statuses import DRAFT, WAIT_PAY, CHECK, WORK
from apps.applications.selectors.course_application import course_application_orm, course_application_model
from apps.applications.selectors.event_application import event_application_orm, event_application_model
from apps.commons.utils.django.exception import exception_handling
from apps.edu.selectors.student_group import student_group_orm


class YearFormsReport:
    """
    Класс для формирования отчета по анкетам групп по типу услуги за год
    """

    # Данные для всех типов услуг
    common_cells = {
        'Телефон': 'profile.phone',
        'Email': 'profile.django_user.email',
        'Фамилия': 'profile.surname',
        'Имя': 'profile.name',
        'Отчество': 'profile.patronymic',
        'Пол': 'special_sex',
        'Регион': 'profile.state.name',
        'Муниципальное образование': 'mo.name',
        'Организация': 'oo.short_name',
        'Категория должности': 'position_category.name',
        'Должность': 'position.name',
        'Тип лица': 'special_physical',
        'Тип ОО': 'oo.oo_type.name',
        'Оплачено': 'special_pay'
    }

    # Данные для курсов
    ou_cells = {
        'Уровень образования': 'education_level',
        'Категория получаемого образования': 'education_category',
        'Фамилия в дипломе': 'diploma_surname',
        'Cерия документа об образовании': 'education_serial',
        'Номер документа об образовании': 'education_number',
        'Дата выдачи документа об образовании': 'special_education_date',
        'Получение удостоверения почтой': 'special_certificate_email',
        'Физический адрес доставки удостоверения': 'mail_address'
    }

    service_type = 'ou'
    year = datetime.date.today().strftime('%Y')

    def __init__(self, report_parameters: dict):
        """
        Инициализация класса - установка типа услуг и года
        :param report_parameters: словарь с параметрами отчета
        """
        self.service_type = report_parameters.get('service_type')
        self.year = report_parameters.get('report_year')

    def _get_groups(self) -> QuerySet:
        """
        Получение учебных групп, подходящих под параметры отчета
        :return: qs учебных групп
        """
        if self.service_type == 'ou':
            return student_group_orm.get_filter_records(
                filter_by=dict(ou__date_start__year=self.year)
            )
        return student_group_orm.get_filter_records(
            filter_by=dict(iku__date_start__year=self.year)
        )

    def _set_columns(self, ws):
        """
        Заполнение заголовков на странице книги Excel
        :param ws: страница книги Excel
        :return:
        """
        index = 1
        for common_cell in self.common_cells.keys():
            if index == 12 and self.service_type == 'ou':
                for ou_index, ou_cell in enumerate(self.ou_cells, start=12):
                    ws.cell(row=1, column=ou_index).value = ou_cell
                index += len(self.ou_cells)
                continue
            ws.cell(row=1, column=index).value = common_cell
            index += 1

    @staticmethod
    def _get_cell_value(app: Union[course_application_model, event_application_model], value: str):
        """
        Выполнение кода из строки (значение в словаре common_cells или ou_cells)
        и попытка получения нужных данных из заявки
        :param value: значение из словаря класса
        :param app: заявка обучающегося
        :return:
        """
        try:
            if value == 'special_sex':
                return 'Муж.' if app.profile.sex else 'Жен.'
            if value == 'special_physical':
                return 'Физ. лицо' if app.physical else 'Юр. лицо'
            elif value == 'special_pay':
                return 'Да' if app.status not in [DRAFT, WORK, WAIT_PAY, CHECK] else 'Нет'
            elif value == 'special_certificate_email':
                return 'Да' if app.certificate_mail else 'Нет'
            elif value == 'education_level':
                return app.get_education_level_display()
            elif value == 'education_category':
                return app.get_education_category_display()
            elif value == 'special_education_date':
                return app.education_date.strftime("%d.%m.%Y")
            else:
                attrs = value.split('.')
                result = app
                for attr in attrs:
                    result = getattr(result, attr)
                return result if result is not None else '-'
        except Exception:
            return '-'

    def _set_data(self, ws):
        """
        Заполнение данных на странице книги Excel
        :param ws: страница книги Excel
        :return:
        """
        row = 2
        for group in self._get_groups():
            if self.service_type == 'ou':
                apps = course_application_orm.get_filter_records(filter_by=dict(group_id=group.object_id))
            else:
                apps = event_application_orm.get_filter_records(filter_by=dict(group_id=group.object_id))
            for app in apps:
                index = 1
                for common_value in self.common_cells.values():
                    if index == 12 and self.service_type == 'ou':
                        for ou_index, ou_value in enumerate(self.ou_cells.values(), start=12):
                            ws.cell(row=row, column=ou_index).value = self._get_cell_value(app, ou_value)
                        index += len(self.ou_cells)
                        continue
                    ws.cell(row=row, column=index).value = self._get_cell_value(app, common_value)
                    index += 1
                row += 1

    def _fill_excel(self, wb: Workbook):
        """
        Заполнение книги Excel
        :param wb: книга Excel
        :return:
        """
        ws = wb.active
        self._set_columns(ws)
        self._set_data(ws)

    def generate_file(self) -> Workbook:
        """Генерация Excel файла с отчетом ПК-1"""
        # Объявление новой книги Excel
        wb = Workbook()
        # Заполнение данными
        self._fill_excel(wb)
        return wb
